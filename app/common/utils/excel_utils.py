"""Excel utility functions."""

from __future__ import annotations

import io

from fastapi.responses import StreamingResponse

from app.core.result import AppError, Err, Ok, Result

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_excel_bytes(
    headers: list[str],
    rows: list[list[object]],
    sheet_title: str = "Sheet",
) -> Result[bytes]:
    """データを Excel ファイル（.xlsx）のバイト列に変換する。

    openpyxl の write_only モードを使用し、メモリ効率よくファイルを生成する。
    大量データの一括エクスポートに対応（10,000 行超はバッチ処理を検討すること）。

    【初学者向け】
    - write_only=True にすると行を追加するだけで読み書き不可になるが、
      メモリを節約できる（一方向エクスポートに最適）。
    - BytesIO はメモリ上の仮想ファイル。ディスクには何も書かない。

    Args:
        headers:     ヘッダー行のセル値リスト。例: ["id", "name", "status"]
        rows:        データ行のリスト。各要素は headers と同順のリスト。
        sheet_title: シート名（Excel 下部に表示される名前）。

    Returns:
        成功時は Ok(bytes)、失敗時は Err(INTERNAL)。

    Example::

        headers = ["申請 ID", "名称", "ステータス"]
        rows = [[a.id, a.name, a.status] for a in applications]
        result = build_excel_bytes(headers, rows, sheet_title="申請一覧")
    """
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_title)
        ws.append(headers)
        for row in rows:
            ws.append(row)

        # メモリ上のバッファに書き込む（ファイルには保存しない）
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)  # 読み取り位置を先頭に戻す
        return Ok(value=output.getvalue())

    except Exception as exc:
        return Err(
            error=AppError(
                type="INTERNAL",
                message="Failed to build Excel file",
                details=str(exc),
            )
        )


def excel_streaming_response(
    content: bytes,
    filename: str,
) -> StreamingResponse:
    """Excel バイト列を HTTP ダウンロードレスポンスに変換する。

    ブラウザは Content-Disposition ヘッダーを受け取ると
    ファイルとして保存するダイアログを表示する。

    Args:
        content:  build_excel_bytes で生成したバイト列。
        filename: ダウンロード時のファイル名。例: "export_2026-04-27.xlsx"

    Returns:
        FastAPI の StreamingResponse（Router の return に直接使用可）。

    Example::

        return excel_streaming_response(excel_bytes, filename="audit-log.xlsx")
    """
    return StreamingResponse(
        iter([content]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def parse_excel_rows(
    file_content: bytes,
    required_columns: set[str] | None = None,
) -> Result[tuple[dict[str, int], list[tuple[object, ...]]]]:
    """Excel ファイルのアクティブシートを解析する。

    1 行目をヘッダーとして扱い、カラム名 → 列インデックスの辞書を返す。
    ヘッダー解決を動的に行うため、列順が変わっても対応可能。

    【初学者向け】
    - read_only=True でメモリ節約（大容量ファイル対応）。
    - data_only=True で数式ではなく計算済みの値を取得。
    - iter_rows(values_only=True) で各行がタプルとして返される。

    Args:
        file_content:     .xlsx ファイルの生バイト列。
        required_columns: 必須カラム名セット（不足時は VALIDATION エラー）。
                          省略時はチェックしない。

    Returns:
        Ok((headers_dict, data_rows)) を返す。
        - headers_dict: カラム名 → 0-based 列インデックス。例: {"id": 0, "name": 1}
        - data_rows:   ヘッダー行を除くデータ行のタプルリスト
        失敗時は Err(VALIDATION) を返す。

    Example::

        result = parse_excel_rows(content, required_columns={"id", "title"})
        if is_err(result):
            return result
        headers, rows = result.value
        for row in rows:
            item_id = row[headers["id"]]
            title   = row[headers["title"]]
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return Err(error=AppError(type="VALIDATION", message="Excel file has no active sheet"))

        rows_iter = ws.iter_rows(values_only=True)

        # 1 行目をヘッダーとして読み込む
        header_row = next(rows_iter, None)
        if header_row is None:
            return Err(error=AppError(type="VALIDATION", message="Excel file is empty"))

        # カラム名 → 列インデックス の辞書を構築する
        headers: dict[str, int] = {}
        for i, cell in enumerate(header_row):
            if cell is not None:
                headers[str(cell).strip()] = i

        # 必須カラムの存在確認
        if required_columns:
            missing = required_columns - headers.keys()
            if missing:
                return Err(
                    error=AppError(
                        type="VALIDATION",
                        message=f"Excel missing required columns: {', '.join(sorted(missing))}",
                    )
                )

        data_rows = list(rows_iter)
        return Ok(value=(headers, data_rows))

    except Exception as exc:
        return Err(
            error=AppError(
                type="VALIDATION",
                message="Failed to parse Excel file",
                details=str(exc),
            )
        )
